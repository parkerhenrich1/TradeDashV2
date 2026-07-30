import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from datetime import datetime

WORKBOOK = 'TradeDash.xlsx'

universe = pd.read_excel(
    "TickerUniverse.xlsx"
)

TICKERS = (
    universe["Symbol"]
    .dropna()
    .tolist()
)

TICKERS = sorted(list(set(TICKERS)))
print(
    f"Loaded {len(TICKERS)} symbols"
)

print(
    "STRL in universe:",
    "STRL" in TICKERS
)

def compute_rsi(series, period=14):
    delta = series.diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)
    avg_gain = gain.rolling(period).mean()
    avg_loss = loss.rolling(period).mean()
    rs = avg_gain / avg_loss
    return 100 - (100 / (1 + rs))


records = []

print('Building TradeDash V3 Universe...')

for symbol in TICKERS:

    try:
        print(f'Scanning {symbol}')

        data = yf.download(
            symbol,
            period='2y',
            interval='1d',
            auto_adjust=True,
            progress=False
        )

        if data.empty:
            continue

        if isinstance(data.columns, pd.MultiIndex):
            data.columns = data.columns.get_level_values(0)

        if len(data) < 252:
            continue

        close = data['Close']
        volume = data['Volume']

        avg_volume = float(
            volume
            .rolling(20)
            .mean()
            .iloc[-1]
        )   

        if avg_volume < 300000:
            continue
        price = float(close.iloc[-1])
        sma20 = float(close.rolling(20).mean().iloc[-1])
        sma50 = float(close.rolling(50).mean().iloc[-1])
        sma200 = float(close.rolling(200).mean().iloc[-1])

        rsi = float(compute_rsi(close).iloc[-1])
        rvol = float(volume.iloc[-1] / volume.rolling(20).mean().iloc[-1])

        return_3m = ((price / close.iloc[-63]) - 1) * 100

        high52 = float(close.rolling(252).max().iloc[-1])
        dist52high = ((price / high52) - 1) * 100

        trend_score = 0
        if price > sma20: trend_score += 10
        if price > sma50: trend_score += 10
        if price > sma200: trend_score += 15
        if sma20 > sma50: trend_score += 10
        if sma50 > sma200: trend_score += 10
        if rsi > 50: trend_score += 5

        rs_score = 40 if return_3m > 50 else 30 if return_3m > 30 else 20 if return_3m > 20 else 10 if return_3m > 10 else 0
        high_score = 30 if dist52high > -2 else 20 if dist52high > -5 else 10 if dist52high > -10 else 0
        volume_score = 30 if rvol > 2 else 20 if rvol > 1.5 else 10 if rvol > 1.2 else 0

        alpha_score = min(trend_score + rs_score + high_score + volume_score, 100)

        entry_score = 0
        if price > sma200: entry_score += 25
        if price > sma50: entry_score += 20
        if 45 <= rsi <= 65: entry_score += 25
        if abs(price - sma20)/sma20 < 0.03: entry_score += 15
        if rvol > 1: entry_score += 15

        elite_leader = (
            alpha_score >= 95 and
            return_3m >= 50 and
            dist52high > -3
        )

        institutional_accum = (
            alpha_score >= 95 and
            rvol >= 3
        )

        decision = 'ELITE LEADER' if elite_leader else 'INSTITUTIONAL ACCUM' if institutional_accum else 'HIGH CONVICTION' if alpha_score >= 95 else 'WATCH' if alpha_score >= 85 else 'AVOID'

        records.append([
            symbol,
            datetime.now().strftime('%Y-%m-%d'),
            round(price,2),
            round(rsi,2),
            round(rvol,2),
            round(return_3m,2),
            round(dist52high,2),
            round(trend_score,2),
            round(entry_score,2),
            round(alpha_score,2),
            elite_leader,
            institutional_accum,
            decision
        ])

    except Exception as e:
        print(symbol, e)

columns = [
    'Symbol','Date','Price','RSI','RVOL','Return3M','Dist52High',
    'TrendScore','EntryScore','AlphaScore',
    'EliteLeader','InstitutionalAccumulation','Decision'
]

indicator_df = pd.DataFrame(records, columns=columns)

ranking_df = indicator_df.sort_values(
    by=['AlphaScore','Return3M','RVOL'],
    ascending=False
).reset_index(drop=True)
ranking_df.insert(0,'Rank',range(1,len(ranking_df)+1))

elite_df = ranking_df[ranking_df['EliteLeader']==True]
inst_df = ranking_df[ranking_df['InstitutionalAccumulation']==True]

universe_df = ranking_df[
    (ranking_df['AlphaScore'] >= 85) |
    (ranking_df['EliteLeader']==True)
]

wb = load_workbook(WORKBOOK)

sheet_map = {
    'Indicators': indicator_df,
    'TradeDashRankings': ranking_df,
    'EliteLeaders': elite_df,
    'InstitutionalAccumulation': inst_df
}

for sheet_name, df in sheet_map.items():
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)

    ws = wb[sheet_name]
    ws.delete_rows(1, ws.max_row)

    for c, h in enumerate(df.columns, start=1):
        ws.cell(1,c,h)

    for r,row in enumerate(df.values.tolist(), start=2):
        for c,val in enumerate(row, start=1):
            ws.cell(r,c,val)

wb.save(WORKBOOK)
print('TradeDash V3 Updated Successfully')
